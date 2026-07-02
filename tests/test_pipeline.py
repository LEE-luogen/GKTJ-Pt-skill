#!/usr/bin/env python3
"""Tests for the GKTJ patient report rendering pipeline."""
from __future__ import annotations

import json
import sys
import zipfile
from io import BytesIO
from pathlib import Path

import pytest
from lxml import etree

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.update_word_charts import (
    ChartSlotData,
    _pct_to_decimal,
    _update_chart_xml,
    _update_workbook,
    payload_to_chart_slots,
)
from scripts.build_payload import (
    _extract_percentages,
    _check_intro_word_count,
    _check_percentages,
    _check_body_percentage_refs,
    text_len,
    validate_payload_v2,
)

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
TEMPLATE_PATH = PROJECT_ROOT / "templates" / "patient-unified-v1.docx"


# ── Unit Tests: Percentage handling ──────────────────────────────────────────

class TestPercentageConversion:
    def test_pct_to_decimal_basic(self):
        assert _pct_to_decimal(25.99) == 0.2599
        assert _pct_to_decimal(100.0) == 1.0
        assert _pct_to_decimal(0.0) == 0.0

    def test_pct_to_decimal_precision(self):
        assert _pct_to_decimal(33.33) == 0.3333
        assert _pct_to_decimal(0.01) == 0.0001

    def test_extract_percentages(self):
        assert _extract_percentages("49%的患者") == [49.0]
        assert _extract_percentages("25.99%和36.5%") == [25.99, 36.5]
        assert _extract_percentages("没有百分比") == []
        assert _extract_percentages("100%和0%") == [100.0, 0.0]


# ── Unit Tests: Text length ──────────────────────────────────────────────────

class TestTextLength:
    def test_basic(self):
        assert text_len("你好世界") == 4

    def test_excludes_whitespace(self):
        assert text_len("你好 世界") == 4
        assert text_len("你好\t世界") == 4
        assert text_len("你好\n世界") == 4

    def test_includes_punctuation(self):
        assert text_len("你好，世界！") == 6


# ── Unit Tests: Payload validation ───────────────────────────────────────────

class TestPayloadValidation:
    def _make_valid_payload(self, num_questions=10):
        items = []
        questions = []
        for i in range(1, num_questions + 1):
            items.append({
                "question_ref": f"q{i:02d}",
                "chart_ref": f"chart_slot_{i:02d}",
                "title": f"测试维度{i}",  # 4 chars
                "analysis": "测试分析文本。" * 50,  # ~300 chars
                "chart": {
                    "options": [
                        {"code": "A", "text": "选项A", "pct": 50.0},
                        {"code": "B", "text": "选项B", "pct": 30.0},
                        {"code": "C", "text": "选项C", "pct": 20.0},
                    ]
                }
            })
            questions.append({
                "number": i,
                "question": f"测试问题{i}",
                "options": [
                    {"code": "A", "text": "选项A"},
                    {"code": "B", "text": "选项B"},
                    {"code": "C", "text": "选项C"},
                ]
            })

        return {
            "meta": {"product": "测试产品", "region": "测试地区"},
            "template": {"template_id": "test", "template_file": "test.docx", "version": "v2"},
            "intro": {
                "paragraphs": ["报告背景文本。" * 50, "数据来源文本。" * 25],
                "background": "报告背景文本。" * 50,  # 300 chars
                "data_source": "数据来源文本。" * 25,  # 175 chars
            },
            "chapter2": {"items": items},
            "feedback": {
                "positive": [{"title": "好", "body": "好的方面。"}],
                "negative": [{"title": "改进", "body": "需改进。"}],
            },
            "summary": {
                "recommendations": [{"title": "建议", "body": "建议内容。"}],
            },
            "attachment": {"questions": questions},
            "checks": {"warnings": [], "errors": []},
        }

    def test_valid_payload_passes(self):
        payload = self._make_valid_payload()
        # Should not raise
        validate_payload_v2(payload)

    def test_eleven_questions_allowed(self):
        payload = self._make_valid_payload(11)
        validate_payload_v2(payload)

    def test_twelve_questions_blocked(self):
        payload = self._make_valid_payload(12)
        with pytest.raises(ValueError, match="maximum is 11"):
            validate_payload_v2(payload)

    def test_missing_required_keys(self):
        payload = {"meta": {}}
        with pytest.raises(ValueError, match="missing required keys"):
            validate_payload_v2(payload)

    def test_background_too_short(self):
        payload = self._make_valid_payload()
        payload["intro"]["background"] = "太短了"
        with pytest.raises(ValueError, match="报告背景字数不足"):
            validate_payload_v2(payload)

    def test_data_source_too_short(self):
        payload = self._make_valid_payload()
        payload["intro"]["data_source"] = "太短"
        with pytest.raises(ValueError, match="数据来源字数不足"):
            validate_payload_v2(payload)

    def test_percentage_out_of_range(self):
        payload = self._make_valid_payload()
        payload["chapter2"]["items"][0]["chart"]["options"][0]["pct"] = 150.0
        with pytest.raises(ValueError, match="百分比超出范围"):
            validate_payload_v2(payload)


# ── Unit Tests: Chart XML update ─────────────────────────────────────────────

class TestChartXmlUpdate:
    def _make_chart_xml(self, series_names, values):
        """Create a minimal chart XML with the given series."""
        nsmap = {
            "c": C_NS,
            "a": A_NS,
        }
        chart = etree.Element(f"{{{C_NS}}}chartSpace", nsmap=nsmap)
        chart_el = etree.SubElement(chart, f"{{{C_NS}}}chart")
        plot_area = etree.SubElement(chart_el, f"{{{C_NS}}}plotArea")
        bar_chart = etree.SubElement(plot_area, f"{{{C_NS}}}barChart")

        for idx, (name, val) in enumerate(zip(series_names, values)):
            ser = etree.SubElement(bar_chart, f"{{{C_NS}}}ser")
            idx_el = etree.SubElement(ser, f"{{{C_NS}}}idx")
            idx_el.set("val", str(idx))

            tx = etree.SubElement(ser, f"{{{C_NS}}}tx")
            str_ref = etree.SubElement(tx, f"{{{C_NS}}}strRef")
            str_cache = etree.SubElement(str_ref, f"{{{C_NS}}}strCache")
            pt = etree.SubElement(str_cache, f"{{{C_NS}}}pt")
            pt.set("idx", "0")
            v = etree.SubElement(pt, f"{{{C_NS}}}v")
            v.text = name

            cat = etree.SubElement(ser, f"{{{C_NS}}}cat")
            cat_str_ref = etree.SubElement(cat, f"{{{C_NS}}}strRef")
            cat_str_cache = etree.SubElement(cat_str_ref, f"{{{C_NS}}}strCache")
            cat_pt = etree.SubElement(cat_str_cache, f"{{{C_NS}}}pt")
            cat_pt.set("idx", "0")
            cat_v = etree.SubElement(cat_pt, f"{{{C_NS}}}v")
            cat_v.text = "占比"

            val_el = etree.SubElement(ser, f"{{{C_NS}}}val")
            num_ref = etree.SubElement(val_el, f"{{{C_NS}}}numRef")
            num_cache = etree.SubElement(num_ref, f"{{{C_NS}}}numCache")
            fmt = etree.SubElement(num_cache, f"{{{C_NS}}}formatCode")
            fmt.text = "0%"
            val_pt = etree.SubElement(num_cache, f"{{{C_NS}}}pt")
            val_pt.set("idx", "0")
            val_v = etree.SubElement(val_pt, f"{{{C_NS}}}v")
            val_v.text = str(val)

        return etree.tostring(chart, xml_declaration=True, encoding="UTF-8", standalone=True)

    def test_update_preserves_series_structure(self):
        chart_bytes = self._make_chart_xml(
            ["A.选项1", "B.选项2", "C.选项3"],
            [0.5, 0.3, 0.2]
        )
        slot = ChartSlotData(
            slot_number=1,
            options=[
                {"code": "A", "text": "新选项1", "pct": 60.0},
                {"code": "B", "text": "新选项2", "pct": 25.0},
                {"code": "C", "text": "新选项3", "pct": 15.0},
            ]
        )
        new_bytes, errors, warnings = _update_chart_xml(chart_bytes, slot)
        assert not errors

        tree = etree.fromstring(new_bytes)
        bar_chart = tree.find(f".//{{{C_NS}}}barChart")
        series = bar_chart.findall(f"{{{C_NS}}}ser")
        assert len(series) == 3

        # Check series names
        for idx, ser in enumerate(series):
            tx_v = ser.find(f"{{{C_NS}}}tx/{{{C_NS}}}strRef/{{{C_NS}}}strCache/{{{C_NS}}}pt/{{{C_NS}}}v")
            assert tx_v.text == f"{slot.options[idx]['code']}.{slot.options[idx]['text']}"

    def test_update_sets_format_code(self):
        chart_bytes = self._make_chart_xml(["A.X"], [0.5])
        slot = ChartSlotData(slot_number=1, options=[{"code": "A", "text": "X", "pct": 50.0}])
        new_bytes, _, _ = _update_chart_xml(chart_bytes, slot)
        tree = etree.fromstring(new_bytes)
        fmt = tree.find(f".//{{{C_NS}}}formatCode")
        assert fmt.text == "0.00%"

    def test_update_sets_values(self):
        chart_bytes = self._make_chart_xml(["A.X", "B.Y"], [0.5, 0.3])
        slot = ChartSlotData(slot_number=1, options=[
            {"code": "A", "text": "X", "pct": 25.99},
            {"code": "B", "text": "Y", "pct": 74.01},
        ])
        new_bytes, _, _ = _update_chart_xml(chart_bytes, slot)
        tree = etree.fromstring(new_bytes)
        val_pts = tree.findall(f".//{{{C_NS}}}val/{{{C_NS}}}numRef/{{{C_NS}}}numCache/{{{C_NS}}}pt/{{{C_NS}}}v")
        assert val_pts[0].text == "0.2599"
        assert val_pts[1].text == "0.7401"

    def test_update_expands_template_series_for_extra_option(self):
        chart_bytes = self._make_chart_xml(["A.X", "B.Y", "C.Z"], [0.5, 0.3, 0.2])
        slot = ChartSlotData(slot_number=1, options=[
            {"code": "A", "text": "X", "pct": 30.0},
            {"code": "B", "text": "Y", "pct": 25.0},
            {"code": "C", "text": "Z", "pct": 20.0},
            {"code": "D", "text": "W", "pct": 25.0},
        ])
        new_bytes, errors, _ = _update_chart_xml(chart_bytes, slot)
        tree = etree.fromstring(new_bytes)
        series = tree.findall(f".//{{{C_NS}}}barChart/{{{C_NS}}}ser")
        assert not errors
        assert len(series) == 4
        assert tree.findall(f".//{{{C_NS}}}ser/{{{C_NS}}}tx//{{{C_NS}}}v")[-1].text == "D.W"


# ── Unit Tests: Workbook update ──────────────────────────────────────────────

class TestWorkbookUpdate:
    def _make_workbook(self, headers, values):
        """Create a minimal workbook with the given data."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="选项")
        for i, h in enumerate(headers, 2):
            ws.cell(row=1, column=i, value=h)
        ws.cell(row=2, column=1, value="占比")
        for i, v in enumerate(values, 2):
            cell = ws.cell(row=2, column=i, value=v)
            cell.number_format = "0%"
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_update_workbook_values(self):
        wb_bytes = self._make_workbook(["A.旧1", "B.旧2"], [0.5, 0.3])
        slot = ChartSlotData(slot_number=1, options=[
            {"code": "A", "text": "新1", "pct": 60.0},
            {"code": "B", "text": "新2", "pct": 40.0},
        ])
        new_bytes, errors, warnings = _update_workbook(wb_bytes, slot)
        assert not errors

        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(new_bytes))
        ws = wb.active
        assert ws.cell(row=1, column=2).value == "A.新1"
        assert ws.cell(row=1, column=3).value == "B.新2"
        assert ws.cell(row=2, column=2).value == 0.6
        assert ws.cell(row=2, column=3).value == 0.4
        assert ws.cell(row=2, column=2).number_format == "0.00%"

    def test_update_workbook_clears_extras(self):
        wb_bytes = self._make_workbook(["A.X", "B.Y", "C.Z"], [0.5, 0.3, 0.2])
        slot = ChartSlotData(slot_number=1, options=[
            {"code": "A", "text": "X", "pct": 100.0},
        ])
        new_bytes, _, _ = _update_workbook(wb_bytes, slot)

        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(new_bytes))
        ws = wb.active
        assert ws.cell(row=1, column=2).value == "A.X"
        assert ws.cell(row=1, column=3).value is None


# ── Unit Tests: Relationship resolution ──────────────────────────────────────

class TestRelationshipResolution:
    def test_template_chart_workbook_mapping(self):
        """Verify chart1.xml maps to Workbook8.xlsx (the key example)."""
        if not TEMPLATE_PATH.exists():
            pytest.skip("Template not found")

        with zipfile.ZipFile(TEMPLATE_PATH) as z:
            doc_rels = etree.fromstring(z.read("word/_rels/document.xml.rels"))
            REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
            CHART_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart"
            PACKAGE_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/package"

            # Find chart1.xml rId
            rid_to_chart = {}
            for rel in doc_rels.findall(f"{{{REL_NS}}}Relationship"):
                if rel.get("Type") == CHART_REL_TYPE:
                    rid_to_chart[rel.get("Id")] = "word/" + rel.get("Target").lstrip("/")

            # Find chart1.xml → workbook mapping
            chart1_path = None
            for rid, path in rid_to_chart.items():
                if "chart1.xml" in path:
                    chart1_path = path
                    break

            assert chart1_path is not None, "chart1.xml not found"

            chart_rels_path = f"word/charts/_rels/chart1.xml.rels"
            chart_rels = etree.fromstring(z.read(chart_rels_path))
            wb_path = None
            for rel in chart_rels.findall(f"{{{REL_NS}}}Relationship"):
                if rel.get("Type") == PACKAGE_REL_TYPE:
                    target = rel.get("Target")
                    resolved = Path("word/charts") / target
                    parts = resolved.parts
                    norm = []
                    for p in parts:
                        if p == "..":
                            if norm: norm.pop()
                        else:
                            norm.append(p)
                    wb_path = "/".join(norm)

            assert wb_path == "word/embeddings/Workbook8.xlsx", f"Expected Workbook8.xlsx, got {wb_path}"


# ── Integration Tests: DOCX structure ────────────────────────────────────────

class TestDocxStructure:
    @pytest.fixture
    def output_docx(self):
        """Generate a test output DOCX."""
        if not TEMPLATE_PATH.exists():
            pytest.skip("Template not found")

        from scripts.render_from_template import render_from_template

        # Build minimal test payload
        items = []
        questions = []
        for i in range(1, 11):
            items.append({
                "question_ref": f"q{i:02d}",
                "chart_ref": f"chart_slot_{i:02d}",
                "title": f"测试维度{i}",  # 4 chars
                "analysis": "测试分析。" * 50,
                "chart": {
                    "options": [
                        {"code": "A", "text": "选项A", "pct": 50.0},
                        {"code": "B", "text": "选项B", "pct": 30.0},
                        {"code": "C", "text": "选项C", "pct": 20.0},
                    ]
                }
            })
            questions.append({
                "number": i,
                "question": f"问题{i}",
                "options": [
                    {"code": "A", "text": "选项A"},
                    {"code": "B", "text": "选项B"},
                    {"code": "C", "text": "选项C"},
                ]
            })

        payload = {
            "meta": {
                "product": "测试产品",
                "disease": "测试疾病",
                "region": "测试地区",
                "report_year": "2025",
                "report_month": "10",
                "company": "测试公司",
            },
            "intro": {
                "background": "报告背景文本。" * 50,  # 300 chars
                "data_source": "数据来源文本。" * 25,  # 175 chars
            },
            "chapter2": {"items": items},
            "feedback": {
                "positive": [{"title": "好", "body": "好的方面。"}],
                "negative": [{"title": "改进", "body": "需改进。"}],
            },
            "summary": {
                "recommendations": [{"title": "建议", "body": "建议内容。"}],
            },
            "attachment": {"questions": questions},
            "checks": {"warnings": [], "errors": []},
        }

        output_path = Path("/tmp/test_gktj_output.docx")
        render_from_template(payload, TEMPLATE_PATH, output_path)
        return output_path

    def test_has_eleven_charts(self, output_docx):
        with zipfile.ZipFile(output_docx) as z:
            charts = [n for n in z.namelist() if n.startswith("word/charts/chart") and n.endswith(".xml")]
            assert len(charts) == 11

    def test_has_eleven_workbooks(self, output_docx):
        with zipfile.ZipFile(output_docx) as z:
            wbs = [n for n in z.namelist() if "Workbook" in n and n.endswith(".xlsx")]
            assert len(wbs) == 11

    def test_has_toc_field(self, output_docx):
        with zipfile.ZipFile(output_docx) as z:
            doc_xml = etree.fromstring(z.read("word/document.xml"))
            fld_chars = doc_xml.findall(f".//{{{W_NS}}}fldChar")
            has_toc = False
            for fc in fld_chars:
                if fc.get(f"{{{W_NS}}}fldCharType") == "begin":
                    parent_r = fc.getparent()
                    next_r = parent_r.getnext()
                    if next_r is not None:
                        instr = next_r.find(f"{{{W_NS}}}instrText")
                        if instr is not None and "TOC" in (instr.text or ""):
                            has_toc = True
            assert has_toc, "No TOC field found"

    def test_update_fields_true(self, output_docx):
        with zipfile.ZipFile(output_docx) as z:
            settings_xml = etree.fromstring(z.read("word/settings.xml"))
            uf = settings_xml.find(f"{{{W_NS}}}updateFields")
            assert uf is not None
            assert uf.get(f"{{{W_NS}}}val") == "true"

    def test_has_page_field_in_footer(self, output_docx):
        with zipfile.ZipFile(output_docx) as z:
            footer_files = [n for n in z.namelist() if "footer" in n.lower() and n.endswith(".xml")]
            has_page = False
            for ff in footer_files:
                footer_xml = etree.fromstring(z.read(ff))
                instr_texts = footer_xml.findall(f".//{{{W_NS}}}instrText")
                for it in instr_texts:
                    if "PAGE" in (it.text or ""):
                        has_page = True
            assert has_page, "No PAGE field in footer"

    def test_has_heading_outline_levels(self, output_docx):
        with zipfile.ZipFile(output_docx) as z:
            doc_xml = etree.fromstring(z.read("word/document.xml"))
            paragraphs = doc_xml.findall(f".//{{{W_NS}}}p")
            outline_levels = set()
            for p in paragraphs:
                ol = p.find(f".//{{{W_NS}}}outlineLvl")
                if ol is not None:
                    outline_levels.add(ol.get(f"{{{W_NS}}}val"))
            assert "0" in outline_levels, "No outline level 0 (chapter headings)"
            assert "1" in outline_levels, "No outline level 1 (section headings)"

    def test_no_png_charts(self, output_docx):
        with zipfile.ZipFile(output_docx) as z:
            png_files = [n for n in z.namelist() if n.endswith(".png")]
            # Allow some PNGs (e.g., cover image) but not chart PNGs
            chart_pngs = [n for n in png_files if "chart" in n.lower()]
            assert len(chart_pngs) == 0, f"Found chart PNGs: {chart_pngs}"


# ── Run ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
