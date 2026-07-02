#!/usr/bin/env python3
"""Update native Word charts and embedded Excel workbooks in a DOCX template.

This module operates directly on OOXML (zip + XML) to update:
1. Chart series names, cache values, and format codes.
2. Embedded Excel workbook cell values and number formats.
3. Chart font family and size (Songti, 10pt).
4. Chart dimensions to prevent text overlap.

It preserves the template's multi-series bar chart structure where each option
is a separate series and the single category is "占比".

Usage as library:
    from update_word_charts import update_charts_in_docx, ChartSlotData

Usage as CLI:
    python update_word_charts.py <docx_path> <payload_v2_json> [-o output.docx]
"""
from __future__ import annotations

import argparse
import copy
import json
import sys
import zipfile
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Optional
from zipfile import ZipFile

import openpyxl
from openpyxl.utils import get_column_letter
from lxml import etree

# ── OOXML namespaces ─────────────────────────────────────────────────────────
NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "wp": "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing",
    "wpsCustomData": "http://www.wps.cn/officeDocument/2013/wpsCustomData",
}
CHART_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart"
PACKAGE_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/package"

# ── Font size constants (hundredths of a point) ──────────────────────────────
CHART_FONT_SIZE_HUNDREDTHS = 1000  # 10pt
MIN_FONT_SIZE_HUNDREDTHS = CHART_FONT_SIZE_HUNDREDTHS  # Backward-compatible API name
FONT_SIZE_LABEL = CHART_FONT_SIZE_HUNDREDTHS
FONT_SIZE_LEGEND = CHART_FONT_SIZE_HUNDREDTHS
FONT_SIZE_AXIS = CHART_FONT_SIZE_HUNDREDTHS
CHART_FONT_NAME = "宋体"


@dataclass
class ChartSlotData:
    """Data for one chart slot to be written into the DOCX."""
    slot_number: int          # 1-based
    options: list[dict]       # [{"code": "A", "text": "1年以内", "pct": 25.99}, ...]
    question_ref: str = ""
    title: str = ""


@dataclass
class ChartSlotResult:
    """Validation result for one chart slot after update."""
    slot_number: int
    chart_xml_path: str
    workbook_path: str
    series_count: int
    expected_series_count: int
    series_names_match: bool
    values_match: bool
    format_ok: bool
    font_size_ok: bool
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return (
            not self.errors
            and self.series_count == self.expected_series_count
            and self.series_names_match
            and self.values_match
            and self.format_ok
            and self.font_size_ok
        )


def _pct_to_decimal(pct: float) -> float:
    """Convert percentage (0-100) to decimal (0-1)."""
    return round(pct / 100.0, 4)


# ── Relationship resolution ──────────────────────────────────────────────────

def _build_rid_to_chart_map(doc_rels_xml: etree._Element) -> dict[str, str]:
    """Map rId → 'word/charts/chartN.xml' from document.xml.rels."""
    result = {}
    for rel in doc_rels_xml.findall("rel:Relationship", NS):
        if rel.get("Type") == CHART_REL_TYPE:
            target = rel.get("Target", "")
            # Normalize: strip leading ../ or /
            path = "word/" + target.lstrip("/")
            result[rel.get("Id")] = path
    return result


def _build_chart_to_wb_map(zf: zipfile.ZipFile, chart_paths: list[str]) -> dict[str, str]:
    """Map chart XML path → workbook path via chart-level .rels files."""
    result = {}
    for chart_path in chart_paths:
        chart_name = Path(chart_path).name
        rels_path = f"word/charts/_rels/{chart_name}.rels"
        wb_path = None
        try:
            rels_xml = etree.fromstring(zf.read(rels_path))
            for rel in rels_xml.findall("rel:Relationship", NS):
                if rel.get("Type") == PACKAGE_REL_TYPE:
                    target = rel.get("Target")
                    # Resolve relative path: word/charts/ + ../embeddings/WorkbookN.xlsx
                    resolved = Path("word/charts") / target
                    # Normalize ../
                    parts = resolved.parts
                    norm = []
                    for p in parts:
                        if p == "..":
                            if norm:
                                norm.pop()
                        else:
                            norm.append(p)
                    wb_path = "/".join(norm)
                    break
        except KeyError:
            pass
        result[chart_path] = wb_path or ""
    return result


def _find_chart_paragraphs(zf: zipfile.ZipFile) -> list[dict]:
    """Find all paragraphs containing chart references in document order.

    Returns list of dicts with keys:
        para_idx: paragraph index in document body
        r_id: relationship ID for the chart
        chart_xml: resolved chart XML path (word/charts/chartN.xml)
        workbook: resolved workbook path (word/embeddings/WorkbookN.xlsx)
    """
    doc_xml = etree.fromstring(zf.read("word/document.xml"))
    doc_rels = etree.fromstring(zf.read("word/_rels/document.xml.rels"))

    rid_to_chart = _build_rid_to_chart_map(doc_rels)
    chart_paths = list(rid_to_chart.values())
    chart_to_wb = _build_chart_to_wb_map(zf, chart_paths)

    paragraphs = doc_xml.findall(".//w:p", NS)
    slots = []
    for i, p in enumerate(paragraphs):
        chart_elem = p.find(".//c:chart", NS)
        if chart_elem is not None:
            r_id = chart_elem.get(f"{{{NS['r']}}}id")
            chart_path = rid_to_chart.get(r_id, "")
            wb_path = chart_to_wb.get(chart_path, "")
            slots.append({
                "para_idx": i,
                "r_id": r_id,
                "chart_xml": chart_path,
                "workbook": wb_path,
            })
    return slots


# ── Chart XML manipulation ───────────────────────────────────────────────────

def _update_chart_xml(
    chart_xml_bytes: bytes,
    slot_data: ChartSlotData,
    min_font_size: int = MIN_FONT_SIZE_HUNDREDTHS,
) -> tuple[bytes, list[str], list[str]]:
    """Update chart XML in-place. Returns (new_bytes, errors, warnings)."""
    errors: list[str] = []
    warnings: list[str] = []

    tree = etree.fromstring(chart_xml_bytes)
    options = slot_data.options

    # Find the barChart element
    bar_chart = tree.find(".//c:chart/c:plotArea/c:barChart", NS)
    if bar_chart is None:
        errors.append("No c:barChart found in chart XML")
        return etree.tostring(tree, xml_declaration=True, encoding="UTF-8", standalone=True), errors, warnings

    # Get existing series
    existing_series = bar_chart.findall("c:ser", NS)

    # The customer template currently seeds three series, while real
    # questionnaires may contain a fourth option. Extend by cloning the last
    # template series so all visual properties are retained.
    if options and not existing_series:
        errors.append("Chart contains no template series to clone")
    while len(existing_series) < len(options) and existing_series:
        clone = copy.deepcopy(existing_series[-1])
        children = list(bar_chart)
        insert_at = max(children.index(series) for series in existing_series) + 1
        bar_chart.insert(insert_at, clone)
        existing_series.append(clone)

    # Update each series
    for idx, option in enumerate(options):
        if idx >= len(existing_series):
            errors.append(f"Option {idx} ({option['code']}) has no matching series in chart")
            continue

        ser = existing_series[idx]
        decimal_val = _pct_to_decimal(option["pct"])

        # Keep series identity and workbook formulas aligned with the option
        # column (B, C, D, E...).
        for tag in ("idx", "order"):
            node = ser.find(f"c:{tag}", NS)
            if node is not None:
                node.set("val", str(idx))
        column = get_column_letter(idx + 2)
        tx_formula = ser.find("c:tx/c:strRef/c:f", NS)
        if tx_formula is not None:
            tx_formula.text = f"Sheet1!${column}$1"
        val_formula = ser.find("c:val/c:numRef/c:f", NS)
        if val_formula is not None:
            val_formula.text = f"Sheet1!${column}$2"

        # Update series name (tx > strRef > strCache > pt > v)
        tx_cache = ser.find("c:tx/c:strRef/c:strCache", NS)
        if tx_cache is not None:
            pt = tx_cache.find("c:pt", NS)
            if pt is not None:
                v = pt.find("c:v", NS)
                if v is not None:
                    v.text = f"{option['code']}.{option['text']}"

        # Update category label (cat > strRef > strCache > pt > v)
        cat_cache = ser.find("c:cat/c:strRef/c:strCache", NS)
        if cat_cache is not None:
            pt = cat_cache.find("c:pt", NS)
            if pt is not None:
                v = pt.find("c:v", NS)
                if v is not None:
                    v.text = "占比"

        # Update value (val > numRef > numCache > pt > v + formatCode)
        val_cache = ser.find("c:val/c:numRef/c:numCache", NS)
        if val_cache is not None:
            # Update format code
            fmt = val_cache.find("c:formatCode", NS)
            if fmt is not None:
                fmt.text = "0.00%"
            # Update value
            pt = val_cache.find("c:pt", NS)
            if pt is not None:
                v = pt.find("c:v", NS)
                if v is not None:
                    v.text = str(decimal_val)

    # If template has more series than options, remove extras
    for idx in range(len(options), len(existing_series)):
        bar_chart.remove(existing_series[idx])
        warnings.append(f"Removed extra series at index {idx}")

    _set_chart_font(tree, min_font_size)

    return etree.tostring(tree, xml_declaration=True, encoding="UTF-8", standalone=True), errors, warnings


def _set_chart_font(chart_xml: etree._Element, font_size: int) -> None:
    """Set every chart text property to Songti at the requested exact size."""
    A_NS = NS["a"]
    C_NS = NS["c"]

    for text_props in list(chart_xml.iter(f"{{{A_NS}}}rPr")) + list(chart_xml.iter(f"{{{A_NS}}}defRPr")) + list(chart_xml.iter(f"{{{A_NS}}}endParaRPr")):
        text_props.set("sz", str(font_size))
        for tag in ("latin", "ea", "cs"):
            font = text_props.find(f"{{{A_NS}}}{tag}")
            if font is None:
                font = etree.SubElement(text_props, f"{{{A_NS}}}{tag}")
            font.set("typeface", CHART_FONT_NAME)

    # Ensure data labels have explicit font size
    for dlbl in chart_xml.iter(f"{{{C_NS}}}dLbl"):
        txpr = dlbl.find(f"{{{C_NS}}}txPr")
        if txpr is None:
            txpr = etree.SubElement(dlbl, f"{{{C_NS}}}txPr")
        # Ensure a:p/a:pPr/a:defRPr exists with correct size
        p_list = txpr.findall(f"{{{A_NS}}}p")
        for p in p_list:
            ppr = p.find(f"{{{A_NS}}}pPr")
            if ppr is None:
                ppr = etree.SubElement(p, f"{{{A_NS}}}pPr")
                p.insert(0, ppr)
            defrpr = ppr.find(f"{{{A_NS}}}defRPr")
            if defrpr is None:
                defrpr = etree.SubElement(ppr, f"{{{A_NS}}}defRPr")
            defrpr.set("sz", str(font_size))
            for tag in ("latin", "ea", "cs"):
                font = defrpr.find(f"{{{A_NS}}}{tag}")
                if font is None:
                    font = etree.SubElement(defrpr, f"{{{A_NS}}}{tag}")
                font.set("typeface", CHART_FONT_NAME)


# ── Workbook manipulation ────────────────────────────────────────────────────

def _update_workbook(
    wb_bytes: bytes,
    slot_data: ChartSlotData,
) -> tuple[bytes, list[str], list[str]]:
    """Update embedded Excel workbook. Returns (new_bytes, errors, warnings)."""
    errors: list[str] = []
    warnings: list[str] = []

    wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    ws = wb.active
    assert ws is not None, "Workbook has no active sheet"
    options = slot_data.options

    # Row 1: header row — ['选项', 'A.xxx', 'B.xxx', 'C.xxx', ...]
    # Row 2: data row   — ['占比', 0.xx, 0.xx, 0.xx, ...]

    # Write header row
    ws.cell(row=1, column=1, value="选项")
    for col_idx, option in enumerate(options, start=2):
        ws.cell(row=1, column=col_idx, value=f"{option['code']}.{option['text']}")

    # Write data row
    ws.cell(row=2, column=1, value="占比")
    for col_idx, option in enumerate(options, start=2):
        decimal_val = _pct_to_decimal(option["pct"])
        cell = ws.cell(row=2, column=col_idx, value=decimal_val)
        cell.number_format = "0.00%"

    # Clear any extra columns beyond our options
    max_col = ws.max_column or 1
    for col_idx in range(len(options) + 2, max_col + 1):
        for row_idx in range(1, min((ws.max_row or 1) + 1, 6)):
            ws.cell(row=row_idx, column=col_idx).value = None

    # Clear extra rows (rows 3+ have dummy data from template)
    for row_idx in range(3, min((ws.max_row or 1) + 1, 10)):
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_idx, column=col_idx).value = None

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), errors, warnings


# ── Main update logic ────────────────────────────────────────────────────────

def update_charts_in_docx(
    docx_path: Path,
    slots_data: list[ChartSlotData],
    output_path: Optional[Path] = None,
    min_font_size: int = MIN_FONT_SIZE_HUNDREDTHS,
) -> tuple[Path, list[ChartSlotResult]]:
    """Update all charts in a DOCX file.

    Args:
        docx_path: Path to the input DOCX.
        slots_data: List of ChartSlotData, one per chart slot (1-indexed).
        output_path: Where to write the output. If None, overwrites input.
        min_font_size: Minimum font size in hundredths of a point.

    Returns:
        (output_path, list of ChartSlotResult)
    """
    if output_path is None:
        output_path = docx_path

    # Read the entire DOCX into memory
    with ZipFile(docx_path, "r") as zf_in:
        file_contents: dict[str, bytes] = {}
        for item in zf_in.namelist():
            file_contents[item] = zf_in.read(item)

    # Find chart paragraphs
    doc_xml = etree.fromstring(file_contents["word/document.xml"])
    doc_rels = etree.fromstring(file_contents["word/_rels/document.xml.rels"])
    rid_to_chart = _build_rid_to_chart_map(doc_rels)
    chart_paths = list(rid_to_chart.values())
    chart_to_wb = _build_chart_to_wb_map_from_contents(file_contents, chart_paths)

    # Find chart-bearing paragraphs
    paragraphs = doc_xml.findall(".//w:p", NS)
    chart_slots_found = []
    for i, p in enumerate(paragraphs):
        chart_elem = p.find(".//c:chart", NS)
        if chart_elem is not None:
            r_id = chart_elem.get(f"{{{NS['r']}}}id")
            chart_path = rid_to_chart.get(r_id, "")
            wb_path = chart_to_wb.get(chart_path, "")
            chart_slots_found.append({
                "para_idx": i,
                "chart_xml": chart_path,
                "workbook": wb_path,
            })

    # Validate slot count
    if len(slots_data) > len(chart_slots_found):
        raise ValueError(
            f"Payload has {len(slots_data)} chart slots but template only has "
            f"{len(chart_slots_found)} chart paragraphs"
        )

    results: list[ChartSlotResult] = []

    # Update each slot
    for slot_idx, slot_data in enumerate(slots_data):
        if slot_idx >= len(chart_slots_found):
            break

        found = chart_slots_found[slot_idx]
        chart_path = found["chart_xml"]
        wb_path = found["workbook"]

        result = ChartSlotResult(
            slot_number=slot_data.slot_number,
            chart_xml_path=chart_path,
            workbook_path=wb_path,
            series_count=0,
            expected_series_count=len(slot_data.options),
            series_names_match=False,
            values_match=False,
            format_ok=False,
            font_size_ok=False,
        )

        # ── Update chart XML ──
        if chart_path and chart_path in file_contents:
            new_chart_bytes, chart_errors, chart_warnings = _update_chart_xml(
                file_contents[chart_path], slot_data, min_font_size
            )
            result.errors.extend(chart_errors)
            result.warnings.extend(chart_warnings)
            file_contents[chart_path] = new_chart_bytes

            # Verify the update
            verify_tree = etree.fromstring(new_chart_bytes)
            bar_chart = verify_tree.find(".//c:chart/c:plotArea/c:barChart", NS)
            if bar_chart is not None:
                series_list = bar_chart.findall("c:ser", NS)
                result.series_count = len(series_list)

                names_ok = True
                values_ok = True
                fmt_ok = True
                for si, ser in enumerate(series_list):
                    if si >= len(slot_data.options):
                        break
                    opt = slot_data.options[si]
                    expected_name = f"{opt['code']}.{opt['text']}"
                    expected_val = str(_pct_to_decimal(opt["pct"]))

                    # Check series name
                    tx_v = ser.find("c:tx/c:strRef/c:strCache/c:pt/c:v", NS)
                    actual_name = tx_v.text if tx_v is not None else ""
                    if actual_name != expected_name:
                        names_ok = False
                        result.errors.append(
                            f"Series {si} name mismatch: expected '{expected_name}', got '{actual_name}'"
                        )

                    # Check value
                    val_v = ser.find("c:val/c:numRef/c:numCache/c:pt/c:v", NS)
                    actual_val = val_v.text if val_v is not None else ""
                    if actual_val != expected_val:
                        values_ok = False
                        result.errors.append(
                            f"Series {si} value mismatch: expected {expected_val}, got {actual_val}"
                        )

                    # Check format code
                    fmt_code = ser.find("c:val/c:numRef/c:numCache/c:formatCode", NS)
                    if fmt_code is not None and fmt_code.text != "0.00%":
                        fmt_ok = False
                        result.errors.append(
                            f"Series {si} format: expected '0.00%', got '{fmt_code.text}'"
                        )

                result.series_names_match = names_ok
                result.values_match = values_ok
                result.format_ok = fmt_ok

                # Check exact chart font contract: Songti 10pt.
                font_ok = True
                font_nodes = (
                    list(verify_tree.iter(f"{{{NS['a']}}}rPr"))
                    + list(verify_tree.iter(f"{{{NS['a']}}}defRPr"))
                    + list(verify_tree.iter(f"{{{NS['a']}}}endParaRPr"))
                )
                for rpr in font_nodes:
                    sz = rpr.get("sz")
                    ea = rpr.find(f"{{{NS['a']}}}ea")
                    if sz != str(min_font_size):
                        font_ok = False
                        result.errors.append(f"Chart font size {sz} does not equal {min_font_size} (10pt)")
                    if ea is None or ea.get("typeface") != CHART_FONT_NAME:
                        font_ok = False
                        result.errors.append("Chart East Asian font is not Songti")
                result.font_size_ok = font_ok
            else:
                result.errors.append("Could not find barChart after update")
        else:
            result.errors.append(f"Chart XML path '{chart_path}' not found in DOCX")

        # ── Update workbook ──
        if wb_path and wb_path in file_contents:
            new_wb_bytes, wb_errors, wb_warnings = _update_workbook(
                file_contents[wb_path], slot_data
            )
            result.errors.extend(wb_errors)
            result.warnings.extend(wb_warnings)
            file_contents[wb_path] = new_wb_bytes
        elif wb_path:
            result.errors.append(f"Workbook path '{wb_path}' not found in DOCX")
        else:
            result.warnings.append("No workbook relationship found for this chart")

        results.append(result)

    # Write the modified DOCX
    _write_docx(output_path, file_contents)

    return output_path, results


def _build_chart_to_wb_map_from_contents(
    file_contents: dict[str, bytes], chart_paths: list[str]
) -> dict[str, str]:
    """Same as _build_chart_to_wb_map but reads from pre-loaded contents."""
    result = {}
    for chart_path in chart_paths:
        chart_name = Path(chart_path).name
        rels_path = f"word/charts/_rels/{chart_name}.rels"
        wb_path = ""
        if rels_path in file_contents:
            rels_xml = etree.fromstring(file_contents[rels_path])
            for rel in rels_xml.findall("rel:Relationship", NS):
                if rel.get("Type") == PACKAGE_REL_TYPE:
                    target = rel.get("Target")
                    resolved = Path("word/charts") / target
                    parts = resolved.parts
                    norm = []
                    for p in parts:
                        if p == "..":
                            if norm:
                                norm.pop()
                        else:
                            norm.append(p)
                    wb_path = "/".join(norm)
                    break
        result[chart_path] = wb_path
    return result


def _write_docx(output_path: Path, file_contents: dict[str, bytes]) -> None:
    """Write modified file contents back to a DOCX (ZIP) file."""
    with ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in file_contents.items():
            zf.writestr(name, data)


# ── Payload → ChartSlotData conversion ───────────────────────────────────────

def payload_to_chart_slots(payload_v2: dict) -> list[ChartSlotData]:
    """Convert a payload v2 dict to a list of ChartSlotData."""
    items = payload_v2.get("chapter2", {}).get("items", [])
    slots = []
    for idx, item in enumerate(items, start=1):
        chart_data = item.get("chart", {})
        # Handle both v2 formats
        options_raw = []
        if "options" in chart_data:
            # New format: chart.options = [{"code", "text", "pct"}]
            options_raw = chart_data["options"]
        elif "series" in chart_data and "categories" in chart_data:
            # Legacy v2 format: chart.series[0].values + chart.categories
            categories = chart_data.get("categories", [])
            values = chart_data.get("series", [{}])[0].get("values", [])
            for ci, cat in enumerate(categories):
                parts = cat.split(".", 1)
                code = parts[0] if parts else ""
                text = parts[1] if len(parts) > 1 else cat
                pct_val = values[ci] * 100 if ci < len(values) and values[ci] <= 1 else (values[ci] if ci < len(values) else 0)
                options_raw.append({"code": code, "text": text, "pct": round(pct_val, 2)})

        slots.append(ChartSlotData(
            slot_number=idx,
            options=options_raw,
            question_ref=item.get("question_ref", f"q{idx:02d}"),
            title=item.get("title", ""),
        ))
    return slots


# ── CLI ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Update native Word charts and embedded workbooks in a DOCX."
    )
    parser.add_argument("docx_path", help="Input DOCX file")
    parser.add_argument("payload_json", help="Payload v2 JSON file")
    parser.add_argument("-o", "--output", help="Output DOCX path (default: overwrite input)")
    parser.add_argument("--min-font-size", type=int, default=1000,
                        help="Chart font size in hundredths of a point (default: 1000 = 10pt)")
    args = parser.parse_args()

    docx_path = Path(args.docx_path)
    payload = json.loads(Path(args.payload_json).read_text(encoding="utf-8"))
    output_path = Path(args.output) if args.output else docx_path

    slots = payload_to_chart_slots(payload)
    out_path, results = update_charts_in_docx(docx_path, slots, output_path, args.min_font_size)

    # Print results
    all_ok = True
    for r in results:
        status = "✓" if r.ok else "✗"
        print(f"  {status} Slot {r.slot_number:02d}: {r.chart_xml_path} → {r.workbook_path}")
        print(f"    Series: {r.series_count}/{r.expected_series_count}, "
              f"names={r.series_names_match}, vals={r.values_match}, "
              f"fmt={r.format_ok}, font={r.font_size_ok}")
        if r.errors:
            all_ok = False
            for e in r.errors:
                print(f"    ERROR: {e}")
        if r.warnings:
            for w in r.warnings:
                print(f"    WARN: {w}")

    if all_ok:
        print(f"\n✓ All {len(results)} chart slots updated successfully → {out_path}")
    else:
        print(f"\n✗ Some chart slots have errors → {out_path}")
        sys.exit(1)


if __name__ == "__main__":
    main()
