#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def normalize_pct(value: object) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        return ""
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        number = float(text)
        if number <= 1:
            number *= 100
        return f"{number:.2f}%"
    except Exception:
        return f"{text}%"


def parse_sheet(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    questions = []
    row = 2
    while row <= ws.max_row:
        q_text = ws.cell(row, 1).value
        if not q_text:
            row += 1
            continue
        count_row = row + 1
        pct_row = row + 2
        options = []
        total_value = ws.cell(count_row, 9).value
        for idx, col in enumerate(range(3, ws.max_column + 1), start=0):
            raw = ws.cell(row, col).value
            if raw in (None, "", "总计"):
                continue
            raw_text = str(raw).strip()
            label = str(raw).split(".", 1)[0].strip()
            if label not in list("ABCDEF"):
                continue
            if "." not in raw_text and raw_text == label:
                continue
            text = raw_text.split(".", 1)[1].strip() if "." in raw_text else raw_text
            count = ws.cell(count_row, col).value
            pct = normalize_pct(ws.cell(pct_row, col).value)
            if text == label and count in (None, "", "None") and pct in ("", "0%", "0.0%", "0.00%"):
                continue
            options.append(
                {
                    "label": label,
                    "text": text,
                    "count": None if count in (None, "") else str(count),
                    "pct": pct,
                }
            )
        questions.append(
            {
                "number": len(questions) + 1,
                "question": str(q_text).strip(),
                "options": options,
                "total": None if total_value in (None, "") else str(total_value),
            }
        )
        row += 3
    return {
        "source_file": str(path),
        "sheet": ws.title,
        "question_count": len(questions),
        "questions": questions,
    }


def main():
    parser = argparse.ArgumentParser(description="Parse questionnaire spreadsheet into normalized JSON.")
    parser.add_argument("input_file", help="Path to xlsx file")
    parser.add_argument("-o", "--output", required=True, help="Output JSON path")
    args = parser.parse_args()

    data = parse_sheet(Path(args.input_file))
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(out)


if __name__ == "__main__":
    main()
